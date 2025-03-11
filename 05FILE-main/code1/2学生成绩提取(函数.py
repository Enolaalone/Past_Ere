import openpyxl as o


def read_excel(name):
    f=open(f'{name}课程.txt','w')
    wb=o.load_workbook('（前3次）2022级导论考试--成绩表.xlsx')
    sheet=wb[f'{name}']
    for row in sheet.iter_rows(max_row=sheet.max_row,min_row=1,max_col=8,min_col=1):#从左到右，从上到下
        strs=''
        n=0
        for cell in row:
            n+=1
            if type(cell.value) is int:#数字
                strs+=str(cell.value)
            if cell.value==None:#缺考
                strs+='\t'
            if type(cell.value) is str:#名字
                if cell.value[-1:]=='*':
                    print(cell.value)
                    strs += cell.value[0:-1]
                else:
                    strs += cell.value
            if n==8:#最后一列
                strs+='\n'
            else:
                strs+='，'
        f.write(strs)
    f.close()

read_excel(598854)
read_excel(598856)
read_excel(598858)
